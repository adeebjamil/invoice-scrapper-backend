from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
from pathlib import Path
import os
import io
import re
import json
import uuid
import base64
import tempfile
import logging

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# ---------------- Mongo ----------------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Bill To Excel API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ---------------- Models ----------------
class ExtractedTable(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    columns: List[str] = Field(default_factory=list)
    rows: List[Dict[str, Any]] = Field(default_factory=list)
    meta: Dict[str, Any] = Field(default_factory=dict)
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class DownloadPayload(BaseModel):
    columns: List[str]
    rows: List[Dict[str, Any]]
    filename: Optional[str] = "bill_data.xlsx"


class BulkDeletePayload(BaseModel):
    ids: Optional[List[str]] = None



# ---------------- Extraction Prompt ----------------
EXTRACTION_PROMPT = """You are a precise invoice / bill / receipt data extractor.

TASK: Extract ONLY the line-item table from this document. Nothing else.

RULES — read carefully:
1. Look at the document and find the main product/service table with rows of items.
2. Use EXACTLY the column headers you see printed in the document (in their original language).
   - If the header is Arabic/Hindi/Urdu etc., write it exactly as-is.
   - Do NOT add extra columns that do not exist in the document.
   - Do NOT create columns like "None", "N/A", or placeholder columns.
3. For each data row, fill in the value you actually see. If a cell is genuinely blank in the document, use an empty string "" — NOT the word "None".
4. Only include columns that have actual data in most rows.
5. The "columns" array must contain only the column names visible in the table header row.

Return ONLY this JSON — no markdown, no explanation, no extra text:

{
  "columns": ["Header1", "Header2", "Header3"],
  "rows": [
    {"Header1": "value", "Header2": "value", "Header3": "value"}
  ],
  "meta": {
    "vendor": "vendor name or empty string",
    "invoice_number": "invoice number or empty string",
    "date": "date or empty string",
    "currency": "currency code or symbol or empty string",
    "language_detected": "english / arabic / hindi / mixed / etc"
  }
}

CRITICAL:
- Return ONLY the JSON object. No ```json fences.
- Every row must have ALL the same keys that are in "columns".
- Missing cell value = "" (empty string), never the word "None".
- Do NOT invent data. Only extract what is physically visible in the document.
- If no table found: {"columns": [], "rows": [], "meta": {"vendor":"","invoice_number":"","date":"","currency":"","language_detected":""}}
"""


def _extract_json_object(text: str) -> Dict[str, Any]:
    """Robustly pull the first {...} JSON object from an LLM response."""
    if not text:
        raise ValueError("Empty response from model")
    # Strip common fences
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
    cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except Exception:
        pass
    # Find largest balanced { ... }
    match = re.search(r"\{[\s\S]*\}", cleaned)
    if not match:
        raise ValueError(f"No JSON object found in model output: {text[:200]}")
    return json.loads(match.group(0))


def _normalize_extraction(data: Dict[str, Any]) -> Dict[str, Any]:
    columns = data.get("columns") or []
    rows = data.get("rows") or []

    # Strip columns that are clearly placeholder/empty
    bad_values = {"none", "n/a", "null", "-", ""}
    
    # Find columns that have at least one real value
    useful_columns = []
    for c in columns:
        col_str = str(c).strip()
        if not col_str or col_str.lower() in bad_values:
            continue
        # Check if this column has any real data across rows
        has_data = any(
            str(r.get(c, "")).strip().lower() not in bad_values
            for r in rows if isinstance(r, dict)
        )
        if has_data:
            useful_columns.append(c)

    # If filtering removed everything, keep originals (better than empty)
    if not useful_columns and columns:
        useful_columns = [c for c in columns if str(c).strip()]

    # Normalize rows — replace "None" / "null" strings with ""
    normalized_rows = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        clean_row = {}
        for c in useful_columns:
            val = str(r.get(c, "")).strip()
            if val.lower() in {"none", "null", "n/a"}:
                val = ""
            clean_row[c] = val
        normalized_rows.append(clean_row)

    return {
        "columns": [str(c) for c in useful_columns],
        "rows": normalized_rows,
        "meta": data.get("meta") or {},
    }


async def _extract_via_gemini(tmp_path: str, mime: str) -> Dict[str, Any]:
    # Kept as a stub — use OpenRouter instead (set OPENROUTER_API_KEY in .env)
    raise HTTPException(500, "Gemini provider disabled. Use OPENROUTER_API_KEY in .env")


async def _extract_via_ollama(file_bytes: bytes, mime: str) -> Dict[str, Any]:
    ollama_url = os.environ.get("OLLAMA_URL", "").rstrip("/")
    model = os.environ.get("OLLAMA_MODEL", "llama3.2-vision")
    if not ollama_url:
        raise HTTPException(500, "OLLAMA_URL not configured")
    if mime == "application/pdf":
        raise HTTPException(400, "Ollama flow supports images only. Send PDF pages as images.")

    b64 = base64.b64encode(file_bytes).decode()
    async with httpx.AsyncClient(timeout=180) as hc:
        resp = await hc.post(
            f"{ollama_url}/api/generate",
            json={
                "model": model,
                "prompt": EXTRACTION_PROMPT,
                "images": [b64],
                "stream": False,
                "format": "json",
            },
        )
        resp.raise_for_status()
        data = resp.json()
        raw = data.get("response", "")
        parsed = _extract_json_object(raw)
        return _normalize_extraction(parsed)


# ---------------- Routes ----------------
@api_router.get("/")
async def root():
    return {"status": "ok", "service": "Bill To Excel"}


@api_router.get("/config")
async def get_config():
    has_openrouter = bool(os.environ.get("OPENROUTER_API_KEY")) and os.environ.get("OPENROUTER_API_KEY") != "sk-or-v1-REPLACE_WITH_YOUR_KEY"
    has_ollama = bool(os.environ.get("OLLAMA_URL"))

    if has_openrouter:
        provider = "openrouter"
        model = os.environ.get("OPENROUTER_MODEL", "google/gemini-2.0-flash-lite-preview-02-05:free")
    elif has_ollama:
        provider = "ollama"
        model = os.environ.get("OLLAMA_MODEL", "llama3.2-vision")
    else:
        provider = "none"
        model = "none"

    return {
        "backend_ready": provider != "none",
        "provider": provider,
        "model": model,
        "openrouter_configured": has_openrouter,
    }


ALLOWED_MIME = {
    "pdf": "application/pdf",
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "webp": "image/webp",
    "heic": "image/heic",
}


async def _extract_via_openrouter(file_bytes: bytes, mime: str) -> Dict[str, Any]:
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key or api_key == "sk-or-v1-REPLACE_WITH_YOUR_KEY":
        raise HTTPException(500, "OPENROUTER_API_KEY not configured. Please set a valid key in .env")

    # Free vision-capable models on OpenRouter — verified & tried in order
    models_to_try = [
        "openrouter/auto",                                    # Auto-selects best free model
        "google/gemma-4-31b-it:free",                        # Google Gemma 4 31B vision (free)
        "google/gemma-4-26b-a4b-it:free",                    # Google Gemma 4 26B vision (free)
        "nvidia/nemotron-nano-12b-v2-vl:free",               # NVIDIA vision model (free)
        "qwen/qwen2.5-vl-32b-instruct:free",                 # Qwen VL 32B (free)
        "qwen/qwen2.5-vl-3b-instruct:free",                  # Qwen VL 3B (free)
        os.environ.get("OPENROUTER_MODEL", "google/gemma-4-31b-it:free"),  # from .env
    ]
    # Deduplicate while preserving order
    seen = set()
    unique_models = []
    for m in models_to_try:
        if m and m not in seen:
            seen.add(m)
            unique_models.append(m)
    models_to_try = unique_models

    b64_data = base64.b64encode(file_bytes).decode("utf-8")
    content_list = [
        {"type": "text", "text": EXTRACTION_PROMPT},
        {
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime};base64,{b64_data}"
            }
        }
    ]

    errors = []
    async with httpx.AsyncClient(timeout=180) as hc:
        for model in models_to_try:
            try:
                logger.info(f"Trying OpenRouter model: {model}")
                resp = await hc.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                        "HTTP-Referer": "https://invoice-scrapper-backend.onrender.com",
                        "X-Title": "Bill To Sheet",
                    },
                    json={
                        "model": model,
                        "messages": [
                            {"role": "user", "content": content_list}
                        ],
                    },
                )
                if resp.status_code == 200:
                    data = resp.json()
                    raw = data["choices"][0]["message"]["content"]
                    logger.info(f"Success with model: {model}")
                    parsed = _extract_json_object(raw)
                    return _normalize_extraction(parsed)
                else:
                    err_msg = f"Model {model} → HTTP {resp.status_code}: {resp.text[:300]}"
                    logger.warning(err_msg)
                    errors.append(err_msg)
            except Exception as e:
                err_msg = f"Model {model} → Exception: {str(e)[:200]}"
                logger.warning(err_msg)
                errors.append(err_msg)

    # All models failed — raise a real error instead of returning fake data
    raise HTTPException(
        502,
        f"All {len(models_to_try)} OpenRouter models failed. "
        f"Check your API key and model availability. "
        f"Last errors: {'; '.join(errors[-3:])}"
    )




@api_router.post("/extract")
async def extract_bill(file: UploadFile = File(...)):
    filename = file.filename or "upload"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    mime = ALLOWED_MIME.get(ext)
    if not mime:
        # Fall back on incoming content_type
        mime = file.content_type or ""
        if mime not in ALLOWED_MIME.values():
            raise HTTPException(400, f"Unsupported file type: {ext or file.content_type}")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Empty file")

    suffix = f".{ext}" if ext else ""
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # Try providers in priority order
        openrouter_key = os.environ.get("OPENROUTER_API_KEY", "")
        has_openrouter = bool(openrouter_key) and openrouter_key != "sk-or-v1-REPLACE_WITH_YOUR_KEY"
        has_ollama = bool(os.environ.get("OLLAMA_URL"))

        if has_openrouter:
            logger.info("Using OpenRouter for extraction")
            result = await _extract_via_openrouter(content, mime)
        elif has_ollama:
            logger.info("Using Ollama for extraction")
            result = await _extract_via_ollama(content, mime)
        else:
            raise HTTPException(
                500,
                "No AI provider configured. Set OPENROUTER_API_KEY (free at openrouter.ai) in backend/.env"
            )

        record = ExtractedTable(
            filename=filename,
            columns=result["columns"],
            rows=result["rows"],
            meta=result.get("meta", {}),
        )
        # Persist history (best-effort, never break the API)
        try:
            await db.extractions.insert_one(record.model_dump())
        except Exception as e:
            logger.warning("Failed to persist extraction: %s", e)

        return {
            "id": record.id,
            "filename": record.filename,
            "columns": record.columns,
            "rows": record.rows,
            "meta": record.meta,
        }
    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        logger.exception("JSON parse failure")
        raise HTTPException(422, f"Model returned invalid JSON: {e}")
    except Exception as e:
        logger.exception("Extraction failed")
        raise HTTPException(500, f"Extraction failed: {e}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


@api_router.post("/download-excel")
async def download_excel(payload: DownloadPayload):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill Data"

    # Detect if content is RTL (Arabic/Hebrew/Urdu)
    def is_rtl(text: str) -> bool:
        return bool(re.search(r'[\u0590-\u05FF\u0600-\u06FF\u0750-\u077F]', str(text)))

    rtl_content = any(is_rtl(c) for c in payload.columns)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="09090B", end_color="09090B", fill_type="solid")
    
    h_align = "right" if rtl_content else "left"
    header_align = Alignment(horizontal=h_align, vertical="center", wrap_text=True, reading_order=2 if rtl_content else 1)
    cell_align  = Alignment(horizontal=h_align, vertical="center", wrap_text=True, reading_order=2 if rtl_content else 1)

    # Write headers
    for i, col in enumerate(payload.columns, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows — try numeric conversion for number-like cells
    for r_idx, row in enumerate(payload.rows, 2):
        for c_idx, col in enumerate(payload.columns, 1):
            raw = row.get(col, "")
            # Try to store as a real number so Excel can sum/sort
            try:
                if raw != "":
                    num = float(str(raw).replace(",", ""))
                    cell = ws.cell(row=r_idx, column=c_idx, value=num)
                else:
                    cell = ws.cell(row=r_idx, column=c_idx, value="")
            except (ValueError, TypeError):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(raw))
            cell.alignment = cell_align

    # Auto-width
    for c_idx, col in enumerate(payload.columns, 1):
        max_len = len(str(col))
        for row in payload.rows:
            v = str(row.get(col, ""))
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(max(14, max_len + 3), 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", payload.filename or "bill_data.xlsx")
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@api_router.get("/history")
async def get_history(limit: int = 50):
    docs = await db.extractions.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs


@api_router.delete("/history/{record_id}")
async def delete_history_item(record_id: str):
    res = await db.extractions.delete_one({"id": record_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "History item not found")
    return {"status": "success", "deleted_id": record_id}


@api_router.delete("/history")
async def delete_history(payload: Optional[BulkDeletePayload] = None):
    if payload and payload.ids:
        res = await db.extractions.delete_many({"id": {"$in": payload.ids}})
    else:
        res = await db.extractions.delete_many({})
    return {"status": "success", "deleted_count": res.deleted_count}



app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
